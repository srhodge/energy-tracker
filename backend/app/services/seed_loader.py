"""
One-time seed loader: reads both Excel sheets and normalizes into the companies table.

Usage:
    python -m app.services.seed_loader path/to/file.xlsx
"""
import html
import sys
import re
from pathlib import Path
from datetime import date

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal, engine
from app.models import Base, Company, Financial, EnergyMaturity, EnergyCategory, EnergySegment, ValueChainPosition


# ---------------------------------------------------------------------------
# WWT model string → energy fields mapping (best-effort defaults)
# ---------------------------------------------------------------------------
MODEL_DEFAULTS: dict[str, dict] = {
    "Chemicals": {
        "energy_category": EnergyCategory.chemicals,
        "energy_segment": EnergySegment.petrochemicals,
        "value_chain_position": ValueChainPosition.downstream,
    },
    "Services": {
        "energy_category": EnergyCategory.energy,
        "energy_segment": EnergySegment.midstream_infrastructure,
        "value_chain_position": ValueChainPosition.services,
    },
    "EPC": {
        "energy_category": EnergyCategory.energy,
        "energy_segment": EnergySegment.midstream_infrastructure,
        "value_chain_position": ValueChainPosition.services,
    },
    "Refining": {
        "energy_category": EnergyCategory.energy,
        "energy_segment": EnergySegment.refined_fuels,
        "value_chain_position": ValueChainPosition.downstream,
    },
    "LNG": {
        "energy_category": EnergyCategory.energy,
        "energy_segment": EnergySegment.integrated_gas,
        "value_chain_position": ValueChainPosition.midstream,
    },
    "Retail": {
        "energy_category": EnergyCategory.energy,
        "energy_segment": EnergySegment.fuel_transport,
        "value_chain_position": ValueChainPosition.downstream,
    },
}


def _clean(val) -> str | None:
    if val is None:
        return None
    s = html.unescape(str(val).strip())
    return s if s else None


def _parse_market_cap(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        # strip suffixes like "$1.2B"
        s = str(val).replace("$", "").replace(",", "").upper().strip()
        mult = {"B": 1e9, "M": 1e6, "T": 1e12, "K": 1e3}
        for suffix, factor in mult.items():
            if s.endswith(suffix):
                try:
                    return float(s[:-1]) * factor
                except ValueError:
                    return None
        try:
            return float(s)
        except ValueError:
            return None


def _rows_from_sheet(ws) -> list[dict]:
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().lower().replace(" ", "_") if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        if all(c is None for c in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def _row_to_company(row: dict) -> dict:
    name = _clean(row.get("name") or row.get("company") or row.get("company_name"))
    if not name:
        return {}

    ticker = _clean(row.get("ticker") or row.get("symbol"))
    country = _clean(row.get("country"))
    territory = _clean(row.get("territory") or row.get("wwt_territory") or row.get("wwt_sales_territory"))
    wwt_model = _clean(row.get("model") or row.get("wwt_model"))
    market_cap = _parse_market_cap(
        row.get("marketcap") or row.get("market_cap") or
        row.get("market_cap_(usd)") or row.get("market_cap_usd")
    )
    price_usd = None
    try:
        price_usd = float(row.get("price_(usd)") or row.get("price_usd") or row.get("price") or 0) or None
    except (TypeError, ValueError):
        pass

    defaults = MODEL_DEFAULTS.get(wwt_model, {})

    return {
        "name": name,
        "ticker": ticker,
        "country": country,
        "wwt_territory": territory,
        "wwt_model": wwt_model,
        "energy_maturity": EnergyMaturity.mature,  # default; can be overridden
        "energy_category": defaults.get("energy_category"),
        "energy_segment": defaults.get("energy_segment"),
        "value_chain_position": defaults.get("value_chain_position"),
        "_market_cap": market_cap,
        "_price_usd": price_usd,
    }


def load_excel(path: str | Path, db: Session) -> int:
    from sqlalchemy import select as _select
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Load names already in DB so re-runs don't duplicate
    existing_names: set[str] = {
        row[0].lower() for row in db.execute(_select(Company.name)).all()
    }
    seen_names: set[str] = set()
    loaded = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name!r}")
        for row in _rows_from_sheet(ws):
            data = _row_to_company(row)
            if not data:
                continue
            key = (data["name"] or "").lower()
            if key in seen_names or key in existing_names:
                continue
            seen_names.add(key)
            existing_names.add(key)

            market_cap = data.pop("_market_cap", None)
            price_usd = data.pop("_price_usd", None)

            company = Company(**data)
            db.add(company)
            db.flush()

            if market_cap or price_usd:
                db.add(Financial(
                    company_id=company.id,
                    market_cap_usd=market_cap,
                    price_usd=price_usd,
                    snapshot_date=date.today(),
                ))

            loaded += 1

    db.commit()
    wb.close()
    return loaded


def main():
    if len(sys.argv) < 2:
        print("Usage: python -m app.services.seed_loader <path_to_excel>")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"File not found: {excel_path}")
        sys.exit(1)

    Base.metadata.create_all(bind=engine)

    with SessionLocal() as db:
        print(f"Loading {excel_path} ...")
        count = load_excel(excel_path, db)
        print(f"Loaded {count} companies.")


if __name__ == "__main__":
    main()
